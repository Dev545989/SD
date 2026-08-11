"""import os
import json
from datetime import datetime, timedelta
from io import BytesIO

import boto3
import openpyxl
from dotenv import load_dotenv

load_dotenv()

# ============================================================
# R2 CONFIG
# ============================================================

R2_ACCESS_KEY = os.getenv("CF_R2_ACCESS_KEY_ID")
R2_SECRET_KEY = os.getenv("CF_R2_SECRET_ACCESS_KEY")
R2_ENDPOINT = os.getenv("CF_R2_ENDPOINT_URL")
BUCKET_NAME = os.getenv("CF_R2_BUCKET_NAME")

s3 = boto3.client(
    "s3",
    endpoint_url=R2_ENDPOINT,
    aws_access_key_id=R2_ACCESS_KEY,
    aws_secret_access_key=R2_SECRET_KEY,
)


# ============================================================
# DATE RANGE
# ============================================================

# 03 → 04
# 04 → 05
# 05 → 06
# 06 → 07
# 07 → 08

START_DATE = datetime(2026, 8, 3)
END_DATE = datetime(2026, 8, 7)


# ============================================================
# LIST OBJECTS
# ============================================================

def list_objects(prefix):

    objects = []

    paginator = s3.get_paginator("list_objects_v2")

    for page in paginator.paginate(
        Bucket=BUCKET_NAME,
        Prefix=prefix,
    ):
        for obj in page.get("Contents", []):
            objects.append(obj["Key"])

    return objects


# ============================================================
# DOWNLOAD
# ============================================================

def download_object(key):

    response = s3.get_object(
        Bucket=BUCKET_NAME,
        Key=key,
    )

    return response["Body"].read()


# ============================================================
# UPLOAD
# ============================================================

def upload_object(key, data, content_type=None):

    kwargs = {
        "Bucket": BUCKET_NAME,
        "Key": key,
        "Body": data,
    }

    if content_type:
        kwargs["ContentType"] = content_type

    s3.put_object(**kwargs)


# ============================================================
# COPY
# ============================================================

def copy_object(source_key, destination_key):

    s3.copy_object(
        Bucket=BUCKET_NAME,
        CopySource={
            "Bucket": BUCKET_NAME,
            "Key": source_key,
        },
        Key=destination_key,
    )


# ============================================================
# DELETE
# ============================================================

def delete_object(key):

    s3.delete_object(
        Bucket=BUCKET_NAME,
        Key=key,
    )


# ============================================================
# REPLACE DATE INSIDE IMAGE PATH
# ============================================================

def replace_date_in_path(
    value,
    old_date,
    new_date,
):

    if not isinstance(value, str):
        return value

    old_prefix = (
        f"DKSA/"
        f"year={old_date.strftime('%Y')}/"
        f"month={old_date.strftime('%m')}/"
        f"day={old_date.strftime('%d')}/"
    )

    new_prefix = (
        f"DKSA/"
        f"year={new_date.strftime('%Y')}/"
        f"month={new_date.strftime('%m')}/"
        f"day={new_date.strftime('%d')}/"
    )

    return value.replace(
        old_prefix,
        new_prefix,
    )


# ============================================================
# UPDATE EXCEL
# ============================================================

def update_excel(
    data,
    old_date,
    new_date,
):

    workbook = openpyxl.load_workbook(
        BytesIO(data),
        read_only=False,
        data_only=False,
    )

    updated_count = 0

    for worksheet in workbook.worksheets:

        headers = {}

        for cell in worksheet[1]:

            if cell.value is not None:

                headers[
                    str(cell.value).strip()
                ] = cell.column

        if "image_r2_paths" not in headers:
            continue

        image_column = headers["image_r2_paths"]

        for row in range(
            2,
            worksheet.max_row + 1,
        ):

            cell = worksheet.cell(
                row=row,
                column=image_column,
            )

            if not isinstance(
                cell.value,
                str,
            ):
                continue

            old_value = cell.value

            new_value = replace_date_in_path(
                old_value,
                old_date,
                new_date,
            )

            if new_value != old_value:

                cell.value = new_value

                updated_count += 1

    output = BytesIO()

    workbook.save(output)

    workbook.close()

    output.seek(0)

    return output.read(), updated_count


# ============================================================
# UPDATE JSON
# ============================================================

def update_json_recursive(
    value,
    old_date,
    new_date,
):

    updated_count = 0

    if isinstance(value, dict):

        for key in value:

            if key == "image_r2_paths":

                if isinstance(
                    value[key],
                    list,
                ):

                    for index, item in enumerate(
                        value[key]
                    ):

                        if isinstance(
                            item,
                            str,
                        ):

                            new_item = replace_date_in_path(
                                item,
                                old_date,
                                new_date,
                            )

                            if new_item != item:
                                updated_count += 1

                            value[key][index] = new_item

                elif isinstance(
                    value[key],
                    str,
                ):

                    old_value = value[key]

                    new_value = replace_date_in_path(
                        old_value,
                        old_date,
                        new_date,
                    )

                    if new_value != old_value:

                        value[key] = new_value

                        updated_count += 1

            else:

                updated_count += update_json_recursive(
                    value[key],
                    old_date,
                    new_date,
                )

    elif isinstance(value, list):

        for item in value:

            if isinstance(
                item,
                (dict, list),
            ):

                updated_count += update_json_recursive(
                    item,
                    old_date,
                    new_date,
                )

    return updated_count


def update_json(
    data,
    old_date,
    new_date,
):

    json_data = json.loads(
        data.decode("utf-8")
    )

    updated_count = update_json_recursive(
        json_data,
        old_date,
        new_date,
    )

    output = json.dumps(
        json_data,
        ensure_ascii=False,
        indent=2,
    ).encode("utf-8")

    return output, updated_count


# ============================================================
# PROCESS FILE
# ============================================================

def process_file(
    source_key,
    destination_key,
    old_date,
    new_date,
):

    extension = source_key.lower().rsplit(
        ".",
        1,
    )[-1]

    # --------------------------------------------------------
    # EXCEL
    # --------------------------------------------------------

    if extension == "xlsx":

        data = download_object(
            source_key
        )

        updated_data, count = update_excel(
            data,
            old_date,
            new_date,
        )

        upload_object(
            destination_key,
            updated_data,
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        print(
            f"      📊 Excel "
            f"image_r2_paths updated: {count}"
        )

    # --------------------------------------------------------
    # JSON
    # --------------------------------------------------------

    elif extension == "json":

        data = download_object(
            source_key
        )

        updated_data, count = update_json(
            data,
            old_date,
            new_date,
        )

        upload_object(
            destination_key,
            updated_data,
            content_type="application/json",
        )

        print(
            f"      📄 JSON "
            f"image_r2_paths updated: {count}"
        )

    # --------------------------------------------------------
    # OTHER FILES
    # --------------------------------------------------------

    else:

        copy_object(
            source_key,
            destination_key,
        )


# ============================================================
# MOVE ONE PREFIX
# ============================================================

def move_prefix(
    source_prefix,
    destination_prefix,
    old_date,
    new_date,
):

    objects = list_objects(
        source_prefix
    )

    if not objects:

        print(
            f"⚠️ NOTHING found under:"
            f"\n   {source_prefix}"
        )

        return

    print()
    print(
        f"📂 FROM: {source_prefix}"
    )
    print(
        f"📂 TO:   {destination_prefix}"
    )
    print(
        f"📦 Objects: {len(objects)}"
    )

    for source_key in objects:

        destination_key = source_key.replace(
            source_prefix,
            destination_prefix,
            1,
        )

        print()
        print(
            f"   ➡️ {source_key}"
        )

        process_file(
            source_key,
            destination_key,
            old_date,
            new_date,
        )

        delete_object(
            source_key
        )

        print(
            "      ✅ Moved"
        )


# ============================================================
# SHIFT ONE DAY
# ============================================================

def shift_date(date):

    new_date = date + timedelta(days=1)

    source_prefix = (
        f"DKSA/"
        f"year={date.strftime('%Y')}/"
        f"month={date.strftime('%m')}/"
        f"day={date.strftime('%d')}/"
    )

    destination_prefix = (
        f"DKSA/"
        f"year={new_date.strftime('%Y')}/"
        f"month={new_date.strftime('%m')}/"
        f"day={new_date.strftime('%d')}/"
    )

    print()
    print("=" * 70)

    print(
        f"📅 {date.strftime('%Y-%m-%d')}"
        f"  →  "
        f"{new_date.strftime('%Y-%m-%d')}"
    )

    print("=" * 70)

    move_prefix(
        source_prefix,
        destination_prefix,
        date,
        new_date,
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 70)
    print("🚀 DKSA R2 Date Shift")
    print("=" * 70)

    print(
        "Moving:"
    )

    print(
        "07 → 08"
    )
    print(
        "06 → 07"
    )
    print(
        "05 → 06"
    )
    print(
        "04 → 05"
    )
    print(
        "03 → 04"
    )

    print()
    print(
        "⚠️ Processing newest → oldest"
    )

    # IMPORTANT:
    # 07 → 08
    # 06 → 07
    # 05 → 06
    # 04 → 05
    # 03 → 04

    current_date = END_DATE

    while current_date >= START_DATE:

        shift_date(
            current_date
        )

        current_date -= timedelta(
            days=1
        )

    print()
    print("=" * 70)
    print(
        "✅ ALL DKSA DATES SHIFTED"
    )
    print("=" * 70)


if __name__ == "__main__":
    main()"""


import os
import json
from datetime import datetime, timedelta
from io import BytesIO

import boto3
import openpyxl
from dotenv import load_dotenv

load_dotenv()

# ============================================================
# R2 CONFIG
# ============================================================

R2_ACCESS_KEY = os.getenv("CF_R2_ACCESS_KEY_ID")
R2_SECRET_KEY = os.getenv("CF_R2_SECRET_ACCESS_KEY")
R2_ENDPOINT = os.getenv("CF_R2_ENDPOINT_URL")
BUCKET_NAME = os.getenv("CF_R2_BUCKET_NAME")

s3 = boto3.client(
    "s3",
    endpoint_url=R2_ENDPOINT,
    aws_access_key_id=R2_ACCESS_KEY,
    aws_secret_access_key=R2_SECRET_KEY,
)


# ============================================================
# REMAINING RANGE
# ============================================================

# Continue only:
#
# 03 → 04
# 02 → 03

START_DATE = datetime(2026, 8, 2)
END_DATE = datetime(2026, 8, 3)


# ============================================================
# LIST OBJECTS
# ============================================================

def list_objects(prefix):

    objects = []

    paginator = s3.get_paginator("list_objects_v2")

    for page in paginator.paginate(
        Bucket=BUCKET_NAME,
        Prefix=prefix,
    ):
        for obj in page.get("Contents", []):
            objects.append(obj["Key"])

    return objects


# ============================================================
# DOWNLOAD
# ============================================================

def download_object(key):

    response = s3.get_object(
        Bucket=BUCKET_NAME,
        Key=key,
    )

    return response["Body"].read()


# ============================================================
# UPLOAD
# ============================================================

def upload_object(key, data, content_type=None):

    kwargs = {
        "Bucket": BUCKET_NAME,
        "Key": key,
        "Body": data,
    }

    if content_type:
        kwargs["ContentType"] = content_type

    s3.put_object(**kwargs)


# ============================================================
# COPY
# ============================================================

def copy_object(source_key, destination_key):

    s3.copy_object(
        Bucket=BUCKET_NAME,
        CopySource={
            "Bucket": BUCKET_NAME,
            "Key": source_key,
        },
        Key=destination_key,
    )


# ============================================================
# DELETE
# ============================================================

def delete_object(key):

    s3.delete_object(
        Bucket=BUCKET_NAME,
        Key=key,
    )


# ============================================================
# REPLACE DATE INSIDE IMAGE PATH
# ============================================================

def replace_date_in_path(
    value,
    old_date,
    new_date,
):

    if not isinstance(value, str):
        return value

    old_prefix = (
        f"DKSA/"
        f"year={old_date.strftime('%Y')}/"
        f"month={old_date.strftime('%m')}/"
        f"day={old_date.strftime('%d')}/"
    )

    new_prefix = (
        f"DKSA/"
        f"year={new_date.strftime('%Y')}/"
        f"month={new_date.strftime('%m')}/"
        f"day={new_date.strftime('%d')}/"
    )

    return value.replace(
        old_prefix,
        new_prefix,
    )


# ============================================================
# UPDATE EXCEL
# ============================================================

def update_excel(
    data,
    old_date,
    new_date,
):

    workbook = openpyxl.load_workbook(
        BytesIO(data),
        read_only=False,
        data_only=False,
    )

    updated_count = 0

    for worksheet in workbook.worksheets:

        headers = {}

        for cell in worksheet[1]:

            if cell.value is not None:

                headers[
                    str(cell.value).strip()
                ] = cell.column

        if "image_r2_paths" not in headers:
            continue

        image_column = headers["image_r2_paths"]

        for row in range(
            2,
            worksheet.max_row + 1,
        ):

            cell = worksheet.cell(
                row=row,
                column=image_column,
            )

            if not isinstance(
                cell.value,
                str,
            ):
                continue

            old_value = cell.value

            new_value = replace_date_in_path(
                old_value,
                old_date,
                new_date,
            )

            if new_value != old_value:

                cell.value = new_value

                updated_count += 1

    output = BytesIO()

    workbook.save(output)

    workbook.close()

    output.seek(0)

    return output.read(), updated_count


# ============================================================
# UPDATE JSON
# ============================================================

def update_json_recursive(
    value,
    old_date,
    new_date,
):

    updated_count = 0

    if isinstance(value, dict):

        for key in value:

            if key == "image_r2_paths":

                if isinstance(
                    value[key],
                    list,
                ):

                    for index, item in enumerate(
                        value[key]
                    ):

                        if isinstance(
                            item,
                            str,
                        ):

                            new_item = replace_date_in_path(
                                item,
                                old_date,
                                new_date,
                            )

                            if new_item != item:
                                updated_count += 1

                            value[key][index] = new_item

                elif isinstance(
                    value[key],
                    str,
                ):

                    old_value = value[key]

                    new_value = replace_date_in_path(
                        old_value,
                        old_date,
                        new_date,
                    )

                    if new_value != old_value:

                        value[key] = new_value

                        updated_count += 1

            else:

                updated_count += update_json_recursive(
                    value[key],
                    old_date,
                    new_date,
                )

    elif isinstance(value, list):

        for item in value:

            if isinstance(
                item,
                (dict, list),
            ):

                updated_count += update_json_recursive(
                    item,
                    old_date,
                    new_date,
                )

    return updated_count


def update_json(
    data,
    old_date,
    new_date,
):

    json_data = json.loads(
        data.decode("utf-8")
    )

    updated_count = update_json_recursive(
        json_data,
        old_date,
        new_date,
    )

    output = json.dumps(
        json_data,
        ensure_ascii=False,
        indent=2,
    ).encode("utf-8")

    return output, updated_count


# ============================================================
# PROCESS FILE
# ============================================================

def process_file(
    source_key,
    destination_key,
    old_date,
    new_date,
):

    extension = source_key.lower().rsplit(
        ".",
        1,
    )[-1]

    # Excel
    if extension == "xlsx":

        data = download_object(
            source_key
        )

        updated_data, count = update_excel(
            data,
            old_date,
            new_date,
        )

        upload_object(
            destination_key,
            updated_data,
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        print(
            f"      📊 Excel updated: "
            f"{count} cell(s)"
        )

    # JSON
    elif extension == "json":

        data = download_object(
            source_key
        )

        updated_data, count = update_json(
            data,
            old_date,
            new_date,
        )

        upload_object(
            destination_key,
            updated_data,
            content_type="application/json",
        )

        print(
            f"      📄 JSON updated: "
            f"{count} path(s)"
        )

    # Images / other files
    else:

        copy_object(
            source_key,
            destination_key,
        )


# ============================================================
# MOVE PREFIX
# ============================================================

def move_prefix(
    source_prefix,
    destination_prefix,
    old_date,
    new_date,
):

    objects = list_objects(
        source_prefix
    )

    if not objects:

        print(
            f"⚠️ NOTHING found under:"
            f"\n   {source_prefix}"
        )

        return

    print()
    print(
        f"📂 FROM: {source_prefix}"
    )
    print(
        f"📂 TO:   {destination_prefix}"
    )
    print(
        f"📦 Remaining objects: {len(objects)}"
    )

    for index, source_key in enumerate(
        objects,
        start=1,
    ):

        destination_key = source_key.replace(
            source_prefix,
            destination_prefix,
            1,
        )

        print()
        print(
            f"[{index}/{len(objects)}]"
        )

        print(
            f"➡️ {source_key}"
        )

        process_file(
            source_key,
            destination_key,
            old_date,
            new_date,
        )

        delete_object(
            source_key
        )

        print(
            "   ✅ Moved"
        )


# ============================================================
# SHIFT ONE DAY
# ============================================================

def shift_date(date):

    new_date = date + timedelta(days=1)

    source_prefix = (
        f"DKSA/"
        f"year={date.strftime('%Y')}/"
        f"month={date.strftime('%m')}/"
        f"day={date.strftime('%d')}/"
    )

    destination_prefix = (
        f"DKSA/"
        f"year={new_date.strftime('%Y')}/"
        f"month={new_date.strftime('%m')}/"
        f"day={new_date.strftime('%d')}/"
    )

    print()
    print("=" * 70)

    print(
        f"📅 {date.strftime('%Y-%m-%d')}"
        f" → "
        f"{new_date.strftime('%Y-%m-%d')}"
    )

    print("=" * 70)

    move_prefix(
        source_prefix,
        destination_prefix,
        date,
        new_date,
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 70)
    print("🚀 CONTINUE DKSA R2 DATE SHIFT")
    print("=" * 70)

    print()
    print("Remaining moves:")
    print("03 → 04")
    print("02 → 03")

    print()
    print("⚠️ Starting from 03 → 04")
    print("⚠️ Only remaining objects will be moved.")

    # --------------------------------------------------------
    # IMPORTANT:
    #
    # We ONLY process:
    #
    # 03 → 04
    # 02 → 03
    #
    # 04, 05, 06, 07 are NOT touched.
    # --------------------------------------------------------

    current_date = END_DATE

    while current_date >= START_DATE:

        shift_date(
            current_date
        )

        current_date -= timedelta(
            days=1
        )

    print()
    print("=" * 70)
    print("✅ REMAINING DKSA DATES COMPLETED")
    print("=" * 70)


if __name__ == "__main__":
    main()

